"""Broker Position Adapter — 券商导出 CSV/Excel 适配"""
import csv, os, json
from pathlib import Path
from datetime import datetime, timezone, timedelta
from typing import Optional

CST = timezone(timedelta(hours=8))

# 标准持仓字段
STANDARD_FIELDS = [
    "symbol", "name", "shares", "available_shares", "cost_price",
    "current_price", "market_value", "weight", "board", "source", "updated_at"
]

# 默认中文字段映射
DEFAULT_FIELD_MAP = {
    "证券代码": "symbol", "股票代码": "symbol", "代码": "symbol",
    "证券名称": "name", "股票名称": "name", "名称": "name",
    "持仓数量": "shares", "持股数量": "shares", "数量": "shares",
    "可用数量": "available_shares", "可用股数": "available_shares",
    "成本价": "cost_price", "持仓成本价": "cost_price",
    "最新价": "current_price", "现价": "current_price", "市价": "current_price",
    "市值": "market_value", "最新市值": "market_value",
    "盈亏": "profit_loss", "持仓盈亏": "profit_loss",
    "盈亏比例": "profit_loss_pct", "盈亏比": "profit_loss_pct",
}


def read_positions(path: str, field_map: dict = None, encoding: str = "auto") -> dict:
    """读取持仓文件 (CSV/Excel), 返回标准化结果"""
    if field_map is None:
        field_map = DEFAULT_FIELD_MAP.copy()

    result = {
        "source_path": path,
        "source_type": "unknown",
        "encoding_used": encoding,
        "field_map": field_map,
        "raw_rows": 0,
        "normalized": [],
        "errors": [],
        "warnings": [],
        "status": "ok",
        "cash": 0.0,
    }

    ext = Path(path).suffix.lower()
    if not os.path.exists(path):
        result["errors"].append(f"文件不存在: {path}")
        result["status"] = "failed"
        return result
    if ext == ".csv":
        _read_csv(path, encoding, field_map, result)
    elif ext in (".xlsx", ".xls"):
        _read_excel(path, field_map, result)
    else:
        result["errors"].append(f"不支持的文件格式: {ext}")
        result["status"] = "failed"
        return result

    # 校验
    _validate(result)
    return result


def _read_csv(path, encoding, field_map, result):
    """读取 CSV"""
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb2312"] if encoding == "auto" else [encoding]
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc) as f:
                reader = csv.DictReader(f)
                rows = list(reader)
            if len(rows) == 0:
                continue
            result["encoding_used"] = enc
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    else:
        result["errors"].append("无法识别编码")
        result["status"] = "failed"
        return

    result["source_type"] = "csv"
    result["raw_rows"] = len(rows)

    for row in rows:
        normalized = {}
        for raw_key, val in row.items():
            std_key = field_map.get(raw_key.strip(), raw_key.strip())
            normalized[std_key] = val.strip() if val else ""
        result["normalized"].append(normalized)


def _read_excel(path, field_map, result):
    """读取 Excel"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        result["source_type"] = "excel"
        result["raw_rows"] = len(rows)
        for row in rows:
            normalized = {}
            for raw_key, val in row.items():
                if raw_key is None:
                    continue
                std_key = field_map.get(str(raw_key).strip(), str(raw_key).strip())
                normalized[std_key] = str(val).strip() if val is not None else ""
            result["normalized"].append(normalized)
    except ImportError:
        result["errors"].append("openpyxl 未安装, 无法读取 Excel")
        result["status"] = "failed"


def _validate(result):
    """校验标准化后的持仓"""
    validated = []
    result["errors"] = []
    result["warnings"] = []

    for i, row in enumerate(result["normalized"]):
        sym = row.get("symbol", "")
        if not sym:
            result["warnings"].append(f"第{i+1}行缺少 symbol")
            continue
        if not sym.startswith(("0", "3", "6", "1", "5", "8", "4")):
            result["warnings"].append(f"{sym}: symbol 格式异常")

        # shares
        try:
            shares = int(float(row.get("shares", 0)))
            if shares % 100 != 0 and sym.upper() != "CASH":
                result["warnings"].append(f"{sym}: {shares}不是100的倍数")
            row["shares"] = shares
        except:
            result["errors"].append(f"{sym}: shares 格式错误")
            continue

        # available_shares
        try:
            avail = int(float(row.get("available_shares", shares)))
            if avail > shares:
                result["warnings"].append(f"{sym}: available_shares>{shares}")
            row["available_shares"] = avail
        except:
            row["available_shares"] = shares

        # prices
        for field in ["cost_price", "current_price"]:
            try:
                row[field] = float(row.get(field, 0))
            except:
                row[field] = 0.0

        # board
        from factor_lab.live.account_profile import get_board
        row["board"] = get_board(sym)
        row["source"] = "broker_adapter"

        # CASH
        if sym.upper() == "CASH":
            try:
                result["cash"] = float(row.get("market_value", row.get("shares", 0)))
            except:
                pass
            continue

        validated.append(row)

    result["normalized"] = validated
    if result["errors"]:
        result["status"] = "failed"
    elif result["warnings"]:
        result["status"] = "partial"


def normalize_to_csv(normalized: list, output_path: str):
    """输出标准化 CSV"""
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=STANDARD_FIELDS + ["profit_loss", "profit_loss_pct"], extrasaction="ignore")
        w.writeheader()
        w.writerows(normalized)
